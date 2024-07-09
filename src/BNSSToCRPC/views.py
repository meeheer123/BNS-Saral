from typing import Union, Optional, Dict
import os
from django.shortcuts import render
from django.http import JsonResponse
from openpyxl import load_workbook

def get_csv_path(filename: str) -> str:
    """
    Constructs a secure file path for a given filename, ensuring it does not contain any directory traversal characters (`..`).
    
    Parameters:
        filename (str): The name of the file for which the path needs to be constructed.
        
    Returns:
        str: The constructed file path as a string. Returns `None` if an error occurs during path construction.
    """
    if '..' in filename:
        raise ValueError('Invalid filename provided')
    
    current_dir = os.path.dirname(os.path.realpath(__file__))
    normalized_path = os.path.normpath(filename)
    
    try:
        return os.path.join(current_dir, '..', '..', 'files', normalized_path)
    except Exception as e:
        print(f"Error occurred while constructing CSV path: {e}")
        return None

def load_bnss_crpc_mapping() -> Dict[str, str]:
    """
    Reads an Excel file containing BNSS to CRPC mappings and constructs a dictionary where BNSS sections are keys
    and CRPC sections are values.

    Returns:
    A dictionary where keys are BNSS sections and values are CRPC sections.
    """
    mapping = {}
    excel_path = get_csv_path('bnss_to_crpc_mapping.xlsx')

    wb = load_workbook(filename=excel_path, read_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 4:
            section_bnss, heading_bnss, section_crpc, heading_crpc = row[:4]
            mapping[str(section_bnss).strip()] = str(section_crpc).strip()
        else:
            print(f"Row with unexpected number of columns: {row}")

    return mapping

def load_bnss_extra_data(bnss: str) -> str:
    """
    Reads an Excel file to find and return the corresponding heading for a given BNSS section.

    Parameters:
        bnss (str): The BNSS section to look up in the Excel file.

    Returns:
        str: The CRPC heading corresponding to the given BNSS section if found, otherwise an empty string.
    """
    excel_path = get_csv_path('bnss_to_crpc_mapping.xlsx')

    try:
        wb = load_workbook(filename=excel_path, read_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            section_bnss, heading_bnss, section_crpc, heading_crpc = [str(cell).strip() if cell is not None else '' for cell in row[:4]]
            if section_bnss == bnss:
                return heading_crpc

    except Exception as e:
        print(f"Error occurred while reading Excel file: {e}")

    return ""

def load_crpc_extra_data(crpc: str) -> str:
    """
    Reads an Excel file to find and return the heading corresponding to a given CRPC section.

    Parameters:
    - crpc: The CRPC section to be searched in the Excel file.

    Returns:
    - The heading corresponding to the given CRPC section if found.
    - An empty string if no match is found or an error occurs.
    """
    excel_path = get_csv_path('bnss_to_crpc_mapping.xlsx')

    try:
        wb = load_workbook(filename=excel_path, read_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            section_bnss, heading_bnss, section_crpc, heading_crpc = [str(cell).strip() if cell is not None else '' for cell in row[:4]]
            
            if section_crpc == crpc:
                return heading_crpc

    except Exception as e:
        print(f"Error occurred while reading Excel file: {e}")

    return ""

def find_crpc_from_bnss(bnss: str) -> Union[str, None]:
    """
    Takes a BNSS section as input and returns the corresponding CRPC section based on a pre-defined mapping loaded from an Excel file.
    
    If the BNSS section maps to a hyphen ('-'), it returns "New Section".
    
    Parameters:
    - bnss: A string representing a BNSS section.
    
    Returns:
    - A string representing the corresponding CRPC section or "New Section" if the mapping is a hyphen ('-').
    """
    mapping = load_bnss_crpc_mapping()
    
    if mapping.get(bnss) == '-':
        return "New Section"
    
    return mapping.get(bnss)

def find_bnss_from_crpc(crpc: str) -> Optional[str]:
    """
    Search for a BNSS section that corresponds to a given CRPC section by loading a mapping from an Excel file and iterating through it.

    Parameters:
    crpc (str): A string representing the CRPC section to be searched.

    Returns:
    Optional[str]: The BNSS section corresponding to the given CRPC section if found, otherwise None.
    """
    mapping = load_bnss_crpc_mapping()

    for bnss, mapped_crpc in mapping.items():
        if mapped_crpc == crpc:
            return bnss
        
    return None

def home(request) -> Union[JsonResponse, render]:
    """
    Handles POST requests to map sections between BNSS and CRPC codes.
    
    Parameters:
        request: The HTTP request object containing method, POST data, etc.
    
    Returns:
        JSON response containing the mapped section and additional data.
        Renders the home page template if the request method is not POST.
    """
    
    if request.method == 'POST':
        section = request.POST.get('section', '').strip()
        code_type = request.POST.get('code_type', '').strip()

        if code_type == 'crpc to bnss':
            bnss_result = find_bnss_from_crpc(section)
            if bnss_result:
                bnss_data = load_bnss_extra_data(bnss_result)
                return JsonResponse({'bnss': bnss_result, 'bnss_data': bnss_data})
            else:
                return JsonResponse({'bnss': 'BNSS section not found for given CRPC.', 'bnss_data': 'BNSS section not found for given CRPC.'})
        elif code_type == 'bnss to crpc':
            crpc_result = find_crpc_from_bnss(section)
            if crpc_result:
                bnss_data = load_bnss_extra_data(section) if crpc_result == "New Section" else load_crpc_extra_data(crpc_result)
                return JsonResponse({'crpc': crpc_result, 'bnss_data': bnss_data})
            else:
                return JsonResponse({'crpc': 'CRPC section not found for given BNSS.', 'bnss_data': 'CRPC section not found for given BNSS.'}, status=404)
        else:
            return JsonResponse({'error': 'Please provide either CRPC or BNSS section number.'}, status=400)

    return render(request, 'BNSSToCRPC/home.html')