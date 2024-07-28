from typing import Union, Optional, Dict
import os
from django.shortcuts import render
from django.http import JsonResponse
from openpyxl import load_workbook

# Function to get path to CSV files
def get_csv_path(filename):
    """
    Return the full path to a CSV file given the filename.

    Parameters:
    filename (str): The name of the CSV file.

    Returns:
    str: The full path to the CSV file.

    Raises:
    ValueError: If the filename contains '..' indicating an invalid filename.
    """
    if '..' in filename:
        raise ValueError('Invalid filename provided')
    
    current_dir = os.path.dirname(os.path.realpath(__file__))
    normalized_path = os.path.normpath(os.path.join(current_dir, 'files', filename))
    
    return normalized_path

def load_bsa_iea_mapping() -> Dict[str, str]:
    """
    Reads an Excel file containing BSA to IEA mappings and constructs a dictionary where BSA sections are keys
    and IEA sections are values.

    Returns:
    A dictionary where keys are BSA sections and values are IEA sections.
    """
    mapping = {}
    excel_path = get_csv_path('bsa_to_iea_mapping.xlsx')

    wb = load_workbook(filename=excel_path, read_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
            section_bsa, section_iea, subject, bsa_data = row[:4]
            mapping[str(section_bsa)] = str(section_iea)

    return mapping

def load_bsa_extra_data(bsa: str) -> str:
    """
    Reads an Excel file to find and return the corresponding heading for a given BSA section.

    Parameters:
        bsa (str): The BSA section to look up in the Excel file.

    Returns:
        str: The IEA heading corresponding to the given BSA section if found, otherwise an empty string.
    """
    excel_path = get_csv_path('bsa_to_iea_mapping.xlsx')

    try:
        wb = load_workbook(filename=excel_path, read_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            section_bsa, section_iea, subject, bsa_data = row[:4]
            if section_bsa == bsa:
                return bsa_data

    except Exception as e:
        print(f"Error occurred while reading Excel file: {e}")

    return ""

def load_iea_extra_data(iea: str) -> str:
    """
    Reads an Excel file to find and return the heading corresponding to a given IEA section.

    Parameters:
    - iea: The IEA section to be searched in the Excel file.

    Returns:
    - The heading corresponding to the given IEA section if found.
    - An empty string if no match is found or an error occurs.
    """
    excel_path = get_csv_path('bsa_to_iea_mapping.xlsx')

    try:
        wb = load_workbook(filename=excel_path, read_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            section_bsa, section_iea, subject, bsa_data = row[:4]
            
            if section_iea == iea:
                return bsa_data

    except Exception as e:
        print(f"Error occurred while reading Excel file: {e}")

    return ""

def find_iea_from_bsa(bsa: str) -> Union[str, None]:
    """
    Takes a BSA section as input and returns the corresponding IEA section based on a pre-defined mapping loaded from an Excel file.
    
    Parameters:
    - bsa: A string representing a BSA section.
    
    Returns:
    - A string representing the corresponding IEA section.
    """
    mapping = load_bsa_iea_mapping()
    return mapping.get(bsa)

def find_bsa_from_iea(iea: str) -> Optional[str]:
    """
    Search for a BSA section that corresponds to a given IEA section by loading a mapping from an Excel file and iterating through it.

    Parameters:
    iea (str): A string representing the IEA section to be searched.

    Returns:
    Optional[str]: The BSA section corresponding to the given IEA section if found, otherwise None.
    """
    mapping = load_bsa_iea_mapping()

    for bsa, mapped_iea in mapping.items():
        if str(mapped_iea).strip() == iea:
            return bsa
        
    return None

def home(request) -> Union[JsonResponse, render]:
    """
    Handles POST requests to map sections between BSA and IEA codes.
    
    Parameters:
        request: The HTTP request object containing method, POST data, etc.
    
    Returns:
        JSON response containing the mapped section and additional data.
        Renders the home page template if the request method is not POST.
    """
    
    if request.method == 'POST':
        section = request.POST.get('section', '').strip()
        code_type = request.POST.get('code_type', '').strip()

        if code_type == 'iea to bsa':
            bsa_result = find_bsa_from_iea(section)
            if bsa_result:
                bsa_data = load_bsa_extra_data(bsa_result)
                return JsonResponse({'bsa': bsa_result, 'bsa_data': bsa_data})
            else:
                return JsonResponse({'bsa': 'BSA section not found for given IEA.', 'bsa_data': 'BSA section not found for given IEA.'})
        elif code_type == 'bsa to iea':
            iea_result = find_iea_from_bsa(section)
            if iea_result:
                bsa_data = load_iea_extra_data(iea_result)
                return JsonResponse({'iea': iea_result, 'bsa_data': bsa_data})
            else:
                return JsonResponse({'iea': 'IEA section not found for given BSA.', 'bsa_data': 'IEA section not found for given BSA.'}, status=404)
        else:
            return JsonResponse({'error': 'Please provide either IEA or BSA section number.'}, status=400)

    return render(request, 'BSAToIEA/home.html')
